import io
import re
from datetime import date
from typing import List

from openpyxl import load_workbook

from .models import RoyaltyRow

MONTH_NAMES = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12,
}

EXPECTED_HEADERS = (
    "Title",
    "Sales Type",
    "Royalty Rate",
    "Display #",
    "ISBN #",
    "Publisher",
    "Partner",
    "Promotion",
    "Sale Territory",
    "Currency",
    "DLP",
    "Price Type",
    "Sales Qty",
    "Revenue",
    "Royalty Earned",
    "Less Distribution Fee",
    "Exchange Rate",
    "Royalty Payable Currency",
    "Royalty Payable",
)


def parse_findaway_report(content: bytes, filename: str) -> List[RoyaltyRow]:
    """
    Parse a Findaway royalty report XLSX file.

    Args:
        content: The raw bytes of the XLSX file.
        filename: The name of the source file (used for source_file field).

    Returns:
        A list of RoyaltyRow objects representing each row in the report.
    """
    wb = load_workbook(io.BytesIO(content))

    # Parse period from Summary sheet
    sale_month = _parse_period(wb["Summary"])

    rows: List[RoyaltyRow] = []

    # Parse data from each sheet
    for sheet_name in ["Retail", "Subscription", "Pool"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        sheet_rows = list(ws.iter_rows(values_only=True))
        if not sheet_rows:
            continue
        # Validate header row
        _validate_headers(sheet_rows[0], sheet_name)
        # Parse data rows (skip header)
        for row in sheet_rows[1:]:
            royalty_row = _parse_row(row, sale_month, filename)
            rows.append(royalty_row)

    return rows


def _validate_headers(header_row: tuple, sheet_name: str) -> None:
    """Validate that the header row matches expected column names."""
    mismatches = []
    for i, (actual, expected) in enumerate(zip(header_row, EXPECTED_HEADERS)):
        if actual != expected:
            mismatches.append(f"column {i}: expected '{expected}', got '{actual}'")

    if len(header_row) != len(EXPECTED_HEADERS):
        mismatches.append(
            f"expected {len(EXPECTED_HEADERS)} columns, got {len(header_row)}"
        )

    if mismatches:
        raise ValueError(
            f"Header mismatch in sheet '{sheet_name}': {'; '.join(mismatches)}"
        )


def _parse_period(summary_sheet) -> date:
    """Extract the sale month from the Summary sheet's Period field."""
    for row in summary_sheet.iter_rows(values_only=True):
        if row[0] == "Period:":
            # Format: "June 01, 2025 - June 30, 2025"
            period_str = row[1]
            match = re.match(r"(\w+) \d+, (\d+)", period_str)
            if match:
                month_name, year = match.groups()
                month = MONTH_NAMES[month_name]
                return date(int(year), month, 1)
    raise ValueError("Could not find Period in Summary sheet")


def _parse_row(row: tuple, sale_month: date, filename: str) -> RoyaltyRow:
    """Parse a single data row into a RoyaltyRow object."""
    return RoyaltyRow(
        sale_month=sale_month,
        source_file=filename,
        title=row[0],
        sales_type=row[1],
        royalty_rate=row[2],
        display_number=row[3],
        isbn=row[4],
        publisher=row[5],
        partner=row[6],
        promotion=row[7],
        sale_territory=row[8],
        currency=row[9],
        dlp=row[10],
        price_type=row[11],
        sales_qty=row[12],
        revenue=row[13],
        royalty_earned=row[14],
        less_distribution_fee=row[15],
        exchange_rate=row[16],
        royalty_payable_currency=row[17],
        royalty_payable=row[18],
    )
