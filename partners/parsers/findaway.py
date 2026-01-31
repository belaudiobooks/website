import io
import re
from datetime import date
from decimal import Decimal
from typing import List

from openpyxl import load_workbook

from partners.models import SaleRecord

MONTH_NAMES = {
    "January": 1,
    "February": 2,
    "March": 3,
    "April": 4,
    "May": 5,
    "June": 6,
    "July": 7,
    "August": 8,
    "September": 9,
    "October": 10,
    "November": 11,
    "December": 12,
}

EXPECTED_HEADERS = (
    "Title",
    "Sales Type",
    "Royalty Rate",
    "Display #",
    "ISBN #",
    "Publisher",
    "Partner",
    "Promotion",
    "Sale Territory",
    "Currency",
    "DLP",
    "Price Type",
    "Sales Qty",
    "Revenue",
    "Royalty Earned",
    "Less Distribution Fee",
    "Exchange Rate",
    "Royalty Payable Currency",
    "Royalty Payable",
)


def parse_findaway_report(
    content: bytes, filename: str, drive_id: str = ""
) -> List[tuple[SaleRecord, str]]:
    """
    Parse a Findaway royalty report XLSX file.

    Args:
        content: The raw bytes of the XLSX file.
        filename: The name of the source file (used for source_file field).
        drive_id: Google Drive file ID (optional, can be added later).

    Returns:
        A list of (SaleRecord, isbn_code) tuples. SaleRecords are unsaved and
        have isbn=None. Caller is responsible for creating ISBN objects and
        assigning them to SaleRecords.

    Raises:
        ValueError: If the sum of parsed row amounts doesn't match the Net Amount.
    """
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        summary_sheet = wb["Summary"]
        # Parse period from Summary sheet
        month_of_sale = _parse_period(summary_sheet)
        # Parse expected net amount for verification
        expected_net_amount = _parse_net_amount(summary_sheet)

        rows_with_isbns: List[tuple[SaleRecord, str]] = []

        # Parse channel names from Summary sheet
        channels = _parse_channels(summary_sheet)

        # Parse data from each channel sheet
        for sheet_name in channels:
            if sheet_name not in wb.sheetnames:
                raise ValueError(
                    f"Channel '{sheet_name}' listed in Summary but sheet not found"
                )
            ws = wb[sheet_name]
            sheet_rows = list(ws.iter_rows(values_only=True))
            if not sheet_rows:
                continue
            # Validate header row
            _validate_headers(sheet_rows[0], sheet_name)
            # Parse data rows (skip header and empty rows)
            for row in sheet_rows[1:]:
                if (
                    not row or row[0] is None
                ):  # Skip empty rows (read_only mode includes them)
                    continue
                sale_record, isbn_code = _parse_row(
                    row, month_of_sale, filename, drive_id
                )
                rows_with_isbns.append((sale_record, isbn_code))

        # Verify total amount matches expected net amount
        rows = [row for row, _ in rows_with_isbns]
        _verify_total_amount(rows, expected_net_amount, filename)

        return rows_with_isbns
    finally:
        wb.close()


def _validate_headers(header_row: tuple, sheet_name: str) -> None:
    """Validate that the header row matches expected column names."""
    mismatches = []
    for i, (actual, expected) in enumerate(zip(header_row, EXPECTED_HEADERS)):
        if actual != expected:
            mismatches.append(f"column {i}: expected '{expected}', got '{actual}'")

    if len(header_row) != len(EXPECTED_HEADERS):
        mismatches.append(
            f"expected {len(EXPECTED_HEADERS)} columns, got {len(header_row)}"
        )

    if mismatches:
        raise ValueError(
            f"Header mismatch in sheet '{sheet_name}': {'; '.join(mismatches)}"
        )


def _parse_period(summary_sheet) -> date:
    """Extract the sale month from the Summary sheet's Period field."""
    for row in summary_sheet.iter_rows(values_only=True):
        if row[0] == "Period:":
            # Format: "June 01, 2025 - June 30, 2025"
            period_str = row[1]
            match = re.match(r"(\w+) \d+, (\d+)", period_str)
            if match:
                month_name, year = match.groups()
                month = MONTH_NAMES[month_name]
                return date(int(year), month, 1)
    raise ValueError("Could not find Period in Summary sheet")


def _parse_net_amount(summary_sheet) -> Decimal:
    """Extract the Net Amount from the Summary sheet."""
    for row in summary_sheet.iter_rows(values_only=True):
        if row[0] == "Net Amount:":
            # Format: "12.28 USD"
            net_amount_str = row[1]
            match = re.match(r"([\d.]+)\s+USD", net_amount_str)
            if match:
                return Decimal(match.group(1))
    raise ValueError("Could not find Net Amount in Summary sheet")


def _parse_channels(summary_sheet) -> List[str]:
    """Extract channel names from the Payments by Channel section."""
    channels = []
    in_channels_section = False
    for row in summary_sheet.iter_rows(values_only=True):
        # Skip empty rows
        if not row or row[0] is None:
            if in_channels_section:
                break
            continue
        if row[0] == "Payments by Channel:":
            in_channels_section = True
            continue
        if in_channels_section:
            # Skip the header row
            if row[0] == "Channel":
                continue
            channels.append(row[0])
    if not channels:
        raise ValueError("Could not find channels in Summary sheet")
    return channels


def _verify_total_amount(
    rows: List[SaleRecord], expected: Decimal, filename: str
) -> None:
    """Verify that the sum of row amounts matches the expected net amount."""
    actual = sum(row.amount for row in rows)
    if actual != expected:
        raise ValueError(
            f"Amount mismatch in '{filename}': "
            f"sum of rows is {actual}, expected Net Amount is {expected}"
        )


def _parse_row(
    row: tuple, month_of_sale: date, filename: str, drive_id: str
) -> tuple[SaleRecord, str]:
    """Parse a single data row into an unsaved SaleRecord object.

    Returns:
        Tuple of (SaleRecord, isbn_code). The SaleRecord.isbn is not set;
        caller must assign it after bulk-creating ISBN objects.
    """
    # Column indices from EXPECTED_HEADERS:
    # 0: Title, 1: Sales Type, 4: ISBN #, 6: Partner, 8: Sale Territory,
    # 12: Sales Qty, 17: Royalty Payable Currency, 18: Royalty Payable
    isbn_code = row[4] or ""
    sale_record = SaleRecord(
        month_of_sale=month_of_sale,
        source_file=filename,
        drive_id=drive_id,
        title=row[0],
        sales_type=row[1],
        retailer=row[6],
        country=row[8],
        quantity=row[12] or 0,
        amount_currency=row[17],
        amount=Decimal(str(row[18])) if row[18] is not None else Decimal("0"),
    )
    return sale_record, isbn_code
